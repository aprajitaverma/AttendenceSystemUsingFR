import face_recognition
import cv2
import openpyxl
from openpyxl.utils import get_column_letter
import datetime

# Get a reference to your webcam
video_capture = cv2.VideoCapture(0)
# 'rtsp://192.168.10.10:554/user=admin_password=tlJwpbo6_channel=1_stream=0.sdp?real_stream'

known_faces_today = []
wb = openpyxl.load_workbook('attendance_reg.xlsx')
sheet = wb.get_sheet_by_name('February')
col = sheet.max_column
col += 1
column = get_column_letter(col)


def write_to_register(n, mode):
    if mode == "mark_present":
        if n == "Unknown":
            pass
        elif n in known_faces_today:
            print("You're already marked present")
            pass
        else:
            known_faces_today.append(n)
            cell_for_date = str(column) + str(1)
            cell_to_write_for_date = sheet[cell_for_date]
            cell_to_write_for_date.value = datetime.datetime.today().strftime('%Y-%m-%d')

            for row in sheet.iter_rows(min_row=1, min_col=1):
                for cell in row:
                    if cell.value:
                        if cell.value == n:
                            row_no_of_this_cell = cell.row
                            cell_no = str(column) + str(row_no_of_this_cell)
                            cell_to_write = sheet[cell_no]
                            cell_to_write.value = "Present"
                            print("You're marked present")

            wb.save('attendance_reg.xlsx')

    elif mode == "mark_absent":
        for cell in sheet[column]:
            if cell.value is None:
                cell.value = "Absent"
        wb.save('attendance_reg.xlsx')


# Load a sample picture and learn how to recognize it.
nishant_image = face_recognition.load_image_file("faces/nishant.jpg")
nishant_face_encoding = face_recognition.face_encodings(nishant_image)[0]

# Load a third sample picture and learn how to recognize it.
aprajita_image = face_recognition.load_image_file("faces/aprajita.jpg")
aprajita_face_encoding = face_recognition.face_encodings(aprajita_image)[0]

# Load a fourth sample picture and learn how to recognize it.
vivek_image = face_recognition.load_image_file("faces/vivek.jpg")
vivek_face_encoding = face_recognition.face_encodings(vivek_image)[0]

# Load a fourth sample picture and learn how to recognize it.
riddhi_image = face_recognition.load_image_file("faces/riddhi.jpg")
riddhi_face_encoding = face_recognition.face_encodings(riddhi_image)[0]

# Load a fourth sample picture and learn how to recognize it.
pankaj_image = face_recognition.load_image_file("faces/pankaj.jpg")
pankaj_face_encoding = face_recognition.face_encodings(pankaj_image)[0]

# Load a fourth sample picture and learn how to recognize it.
aekanshu_image = face_recognition.load_image_file("faces/aekanshu.jpg")
aekanshu_face_encoding = face_recognition.face_encodings(aekanshu_image)[0]

# Create arrays of known face encodings and their names
known_face_encodings = [
    nishant_face_encoding,
    aprajita_face_encoding,
    vivek_face_encoding,
    riddhi_face_encoding,
    pankaj_face_encoding,
    aekanshu_face_encoding
]
known_face_names = [
    "Nishant Sambyal",
    "Aprajita Verma",
    "Vivek",
    "Riddhi Singh",
    "Pankaj Negi",
    "Aekanshu Panchal"
    
]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

while True:
    # Grab a frame of the video
    ret, frame = video_capture.read()

    # Resize frame of video to smaller size for faster  processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]

            face_names.append(name)
            write_to_register(name, "mark_present")

    process_this_frame = not process_this_frame

    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

    # Display the resulting image
    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        video_capture.release()
        break

write_to_register(n=None, mode="mark_absent")
# Release handle to the web camera
video_capture.release()
cv2.destroyAllWindows()
